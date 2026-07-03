"""
导出器 - 将测试用例导出为 Excel/CSV/JSON 格式
"""

from __future__ import annotations

import csv
import json
import io
from datetime import datetime
from typing import Optional

from .models import TestProject


def export_to_excel(project: TestProject) -> bytes:
    """导出为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请先安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 定义样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="微软雅黑", size=10)
    body_align = Alignment(vertical="top", wrap_text=True)

    priority_colors = {
        "高": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "中": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 标题行
    headers = ["用例ID", "模块", "标题", "前置条件", "测试步骤", "预期结果",
               "优先级", "测试类型", "标签", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    row_idx = 2
    for scenario in project.scenarios:
        for tc in scenario.test_cases:
            values = [
                tc.id,
                tc.module,
                tc.title,
                tc.precondition,
                tc.steps,
                tc.expected_result,
                tc.priority,
                tc.test_type,
                tc.tags,
                tc.notes,
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value or "")
                cell.font = body_font
                cell.alignment = body_align
                cell.border = thin_border
                if col == 7 and value in priority_colors:
                    cell.fill = priority_colors[value]
            row_idx += 1

    # 自动调整列宽
    column_widths = {
        1: 14, 2: 16, 3: 30, 4: 20, 5: 35, 6: 30,
        7: 10, 8: 14, 9: 16, 10: 20,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加统计sheet
    ws_stats = wb.create_sheet(title="统计概览")
    ws_stats.title = "统计"
    stats_headers = ["指标", "数值"]
    for col, h in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    stats_data = [
        ("项目名称", project.name),
        ("测试框架", project.target_framework),
        ("功能模块数", project.total_scenarios),
        ("总测试用例数", project.total_test_cases),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("各优先级分布", ""),
        ("高优先级", sum(1 for s in project.scenarios for tc in s.test_cases if tc.priority == "高")),
        ("中优先级", sum(1 for s in project.scenarios for tc in s.test_cases if tc.priority == "中")),
        ("低优先级", sum(1 for s in project.scenarios for tc in s.test_cases if tc.priority == "低")),
        ("", ""),
        ("各测试类型分布", ""),
    ]

    type_counts: dict[str, int] = {}
    for s in project.scenarios:
        for tc in s.test_cases:
            type_counts[tc.test_type] = type_counts.get(tc.test_type, 0) + 1
    for t, c in type_counts.items():
        stats_data.append((t, c))

    for i, (key, val) in enumerate(stats_data, 2):
        cell_k = ws_stats.cell(row=i, column=1, value=key)
        cell_v = ws_stats.cell(row=i, column=2, value=val)
        cell_k.font = body_font
        cell_v.font = body_font
        cell_k.border = thin_border
        cell_v.border = thin_border

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_csv(project: TestProject) -> str:
    """导出为 CSV 文件"""
    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["用例ID", "模块", "标题", "前置条件", "测试步骤", "预期结果",
               "优先级", "测试类型", "标签", "备注"]
    writer.writerow(headers)

    for scenario in project.scenarios:
        for tc in scenario.test_cases:
            writer.writerow([
                tc.id, tc.module, tc.title, tc.precondition, tc.steps,
                tc.expected_result, tc.priority, tc.test_type, tc.tags, tc.notes,
            ])

    return output.getvalue()


def export_to_json(project: TestProject) -> str:
    """导出为 JSON 文件"""
    data = {
        "project": {
            "name": project.name,
            "description": project.description,
            "version": project.version,
            "framework": project.target_framework,
            "generated_at": datetime.now().isoformat(),
        },
        "statistics": {
            "total_scenarios": project.total_scenarios,
            "total_test_cases": project.total_test_cases,
        },
        "scenarios": [],
    }

    for scenario in project.scenarios:
        scenario_data = {
            "name": scenario.name,
            "description": scenario.description,
            "method": scenario.method,
            "test_cases": [],
        }
        for tc in scenario.test_cases:
            scenario_data["test_cases"].append(tc.to_dict())
        data["scenarios"].append(scenario_data)

    return json.dumps(data, ensure_ascii=False, indent=2)


def export_project(project: TestProject, filepath: str) -> str:
    """导出项目（根据扩展名自动选择格式）"""
    ext = filepath.lower().split(".")[-1]

    if ext == "xlsx":
        data = export_to_excel(project)
        with open(filepath, "wb") as f:
            f.write(data)
        return "excel"
    elif ext == "csv":
        data = export_to_csv(project)
        with open(filepath, "w", encoding="utf-8-sig") as f:
            f.write(data)
        return "csv"
    elif ext == "json":
        data = export_to_json(project)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(data)
        return "json"
    elif ext in ("py", "java"):
        code = generate_code(project, project.target_framework)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(code)
        return "code"
    else:
        data = project.to_dict()
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return "project"
